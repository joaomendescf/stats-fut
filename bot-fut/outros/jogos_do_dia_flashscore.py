# !pip install selenium
# !apt-get update
# !apt install chromiun-chromedriver

from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook, Workbook
from time import sleep
from tqdm import tqdm
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import os


def abrir_pagina_flashscore():
	options = Options()

	# options.add_argument('window-size=300,600')
	options.add_argument('--headless')
	# options.add_argument('--maximize')
	# options.add_argument('--no-sandbox')
	# options.add_argument('--disable-dev-shm-usage')

	navegador = webdriver.Chrome('chromedriver',chrome_options=options)
	navegador.get("https://www.flashscore.com/")
	# navegador.maximize_window()
	
	#Fechando botao de cookies
	try:
			button_cookies = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button#onetrust-accept-btn-handler')))
			# button_cookies = navegador.find_element(By.CSS_SELECTOR,'button#onetrust-accept-btn-handler')
			button_cookies.click()
	except:
			pass

	botao_todos = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.filters__tab')))

	return navegador

def coletar_id_jogos(navegador, periodo):

	if periodo == 'amanha':
		navegador.find_element(By.CSS_SELECTOR,'button.calendar__navigation--tomorrow').click()
		sleep(3)

	#Pegando id dos jogos
	id_jogos = []
	jogos = navegador.find_elements(By.CSS_SELECTOR,'div.event__match--scheduled')
	for i in jogos:
			try:
					id_jogos.append(i.get_attribute("id"))
			except:
					pass
	id_jogos = [i[4:] for i in id_jogos]

	return id_jogos  

def salvar_id_jogos_txt(id_jogos):
	with open ('arquivo-id.txt', 'w', encoding='utf-8') as arquivo:
		for id in id_jogos:
			arquivo.write(f'{id}\n')

def coletar_dados_gerais_partida(navegador, link):
	navegador.get(f'https://www.flashscore.com/match/{link}/#/match-summary')
	sleep(2)

	try:
		#Data
		data = navegador.find_element(By.CSS_SELECTOR,'div.duelParticipant__startTime').text.split(' ')[0]
		#Horario
		hora = navegador.find_element(By.CSS_SELECTOR,'div.duelParticipant__startTime').text.split(' ')[1]
		#Pais
		pais = navegador.find_element(By.CSS_SELECTOR,'span.tournamentHeader__country').text.split(':')[0]
		#Liga
		liga = navegador.find_element(By.CSS_SELECTOR,'span.tournamentHeader__country')
		liga = liga.find_element(By.CSS_SELECTOR,'a').text
		#home
		home = navegador.find_element(By.CSS_SELECTOR,'div.duelParticipant__home')
		home = home.find_element(By.CSS_SELECTOR,'div.participant__participantName').text
		home = home.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('â','a').replace('ê','e').replace('î','i').replace('ô','o').replace('û','u').replace('à','a').replace('è','e').replace('ì','i').replace('ò','o').replace('ù','u').replace('','').replace('ç','c').replace('ã','a').replace('õ','o')
		home = home.upper()
		#away
		away = navegador.find_element(By.CSS_SELECTOR,'div.duelParticipant__away')
		away = away.find_element(By.CSS_SELECTOR,'div.participant__participantName').text
		away = away.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('â','a').replace('ê','e').replace('î','i').replace('ô','o').replace('û','u').replace('à','a').replace('è','e').replace('ì','i').replace('ò','o').replace('ù','u').replace('','').replace('ç','c').replace('ã','a').replace('õ','o')
		away = away.upper()

	except:
		data = '-'
		hora = '-'
		pais = '-'
		liga = '-'
		home = '-'
		away = '-'

	return data, hora, pais, liga, home, away

def coletar_odds_pre_live_partida(navegador, link):

	jogo_odds = {'odd_home':[],'odd_draw':[],'odd_away':[]}
	navegador.get(f'https://www.flashscore.com/match/{link}/#/odds-comparison/1x2-odds/full-time')
	sleep(2)

	try:	
		######### Acessando Match Odds
		celulas = navegador.find_elements(By.CSS_SELECTOR,'div.ui-table__row')
		odd_home = 0
		odd_draw = 0
		odd_away = 0

		if 'title="bet365"' in str(navegador.find_element(By.CSS_SELECTOR,'div.ui-table.oddsCell__odds')):
				for celula in celulas:
						bookie = celula.find_element(By.CSS_SELECTOR,'img.prematchLogo')
						bookie = bookie.get_attribute('title')
						if ((bookie == 'bet365') & (odd_home == 0)) | ((bookie == 'Betfair') & (odd_home == 0)):
								odd_home = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[0].text
								odd_draw = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[1].text 
								odd_away = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[2].text
						else:
								pass
		else:
				for celula in celulas:
						odd_home = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[0].text
						odd_draw = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[1].text 
						odd_away = celula.find_elements(By.CSS_SELECTOR,'a.oddsCell__odd')[2].text

	except:
		odd_home = '-'
		odd_draw = '-' 
		odd_away = '-'
					
	return odd_away, odd_home, odd_draw

def coletar_classificacao_campeonato(navegador, link, time_casa, time_fora):	

	posicoes = {'pos_time_casa':[],'pos_time_fora':[]}
	navegador.get(f'https://www.flashscore.com.br/jogo/{link}/#/classificacao/table/overall')
	sleep(2)
		
	try:	
		conteudo = navegador.page_source
		site = BeautifulSoup(conteudo, 'html.parser')

		dados_campeonato = site.findAll('div', attrs={'class':'ui-table__row'})

		for dado in dados_campeonato:
			pos = dado.find('div', attrs={'class':'tableCellRank'}).text
			pos = pos.replace('.', '')
			time = dado.find('a', attrs={'class':'tableCellParticipant__name'}).text
			if time == time_casa:
				pos_time_casa = int(pos)
			elif time == time_fora:
				pos_time_fora = int(pos)

		posicoes['pos_time_casa'] = pos_time_casa
		posicoes['pos_time_fora'] = pos_time_fora

	except:
		posicoes['pos_time_casa'] = '-'
		posicoes['pos_time_fora'] = '-'
		
	return posicoes
	
def coletar_dados_h2h(navegador, link, time_A, time_B):
	
	#BUSCANDO RESULTADOS ANTERIORES
	navegador.get(f'https://www.flashscore.com.br/jogo/{link}/#/h2h/overall')
	sleep(2)

	#clicar botao aceitar cookies
	try:
		button_cookies = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'button#onetrust-accept-btn-handler')))
		button_cookies.click()
	except:
		pass
	
	#mostrar mais partidas
	for i in range(3):
		try:
			mostar_mais_casa = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="detail"]/div[7]/div[2]/div[1]/div[3]')))
			mostar_mais_casa.click()
			sleep(1.5)
		except:
			pass

	for i in range(3):
		try:
			mostar_mais_fora = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="detail"]/div[7]/div[2]/div[2]/div[3]')))
			mostar_mais_fora.click()
			sleep(1.5)
		except:
			pass
		
	for i in range(3):
		try:
			mostar_mais_casa = WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="detail"]/div[7]/div[2]/div[3]/div[3]')))
			mostar_mais_casa.click()
			sleep(1.5)
		except:
			pass

	#dados dos jogos
	conteudo = navegador.page_source
	site = BeautifulSoup(conteudo, 'html.parser')

	resultados_times = site.findAll('div', attrs={'class':'h2h__row'})

	qtd_confronto_direto = 0
	
	#dados time A
	time_A_casa_qtd = 0
	time_A_casa_v = 0
	time_A_casa_d = 0
	time_A_casa_e = 0
	time_A_casa_gm = 0
	time_A_casa_gs = 0
	time_A_fora_qtd = 0
	time_A_fora_v = 0
	time_A_fora_d = 0
	time_A_fora_e = 0
	time_A_fora_gm = 0
	time_A_fora_gs = 0
	time_A_qtd_mais_35 = 0
	time_A_qtd_mais_25 = 0
	time_A_qtd_mais_15 = 0
	time_A_qtd_mais_05 = 0

	#dados time B
	time_B_casa_qtd = 0
	time_B_casa_v = 0
	time_B_casa_d = 0
	time_B_casa_e = 0
	time_B_casa_gm = 0
	time_B_casa_gs = 0
	time_B_fora_qtd = 0
	time_B_fora_v = 0
	time_B_fora_d = 0
	time_B_fora_e = 0
	time_B_fora_gm = 0
	time_B_fora_gs = 0
	time_B_qtd_mais_35 = 0
	time_B_qtd_mais_25 = 0
	time_B_qtd_mais_15 = 0
	time_B_qtd_mais_05 = 0
	
	cd_casa_v = 0
	cd_casa_d = 0
	cd_casa_e = 0
	cd_casa_gm = 0
	cd_casa_gs = 0
	status = ''
	texto = ''

	total_jogos_time_A = 1
	total_jogos_time_B = 1
	
	for resultado in resultados_times:
		
		try:
			data = resultado.find('span', attrs={'class':'h2h__date'})
			data = data.text

			campeonato = resultado.find('span', attrs={'class':'h2h__event'})
			campeonato = campeonato['title']

			time_casa = resultado.find('span', attrs={'class':'h2h__homeParticipant'})
			time_casa = time_casa.find('span', attrs={'class':'h2h__participantInner'}).text
			time_casa = time_casa.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('â','a').replace('ê','e').replace('î','i').replace('ô','o').replace('û','u').replace('à','a').replace('è','e').replace('ì','i').replace('ò','o').replace('ù','u').replace('','').replace('ç','c').replace('ã','a').replace('õ','o')
			time_casa = time_casa.upper()

			time_fora = resultado.find('span', attrs={'class':'h2h__awayParticipant'})
			time_fora = time_fora.find('span', attrs={'class':'h2h__participantInner'}).text
			time_fora = time_fora.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u').replace('â','a').replace('ê','e').replace('î','i').replace('ô','o').replace('û','u').replace('à','a').replace('è','e').replace('ì','i').replace('ò','o').replace('ù','u').replace('','').replace('ç','c').replace('ã','a').replace('õ','o')
			time_fora = time_fora.upper()
			
			placar = resultado.find('span', attrs={'class':'h2h__result'})
			placar = placar.findAll('span')
			placar_casa = placar[0].text
			placar_fora = placar[1].text

			# if (time_A == time_casa or time_A == time_fora) and (time_B == time_casa or time_B == time_fora):
			# 	qtd_confronto_direto += 1
			
			if (time_A == time_casa) or (time_A == time_fora):
				print(f'{time_casa} == {time_fora}')
				if total_jogos_time_A <= 15:
					if time_A == time_casa:
						if int(placar_casa) > int(placar_fora):
							time_A_casa_v += 1
							status = 'VC'
						elif int(placar_casa) < int(placar_fora):
							time_A_casa_d += 1
							status = 'DC'				
						elif int(placar_casa) == int(placar_fora):
							time_A_casa_e += 1
							status = 'EC'

						time_A_casa_qtd += 1
						time_A_casa_gm += int(placar_casa)
						time_A_casa_gs += int(placar_fora)
					
					elif time_A == time_fora:
						if int(placar_casa) > int(placar_fora):
							time_A_fora_d += 1
							status = 'DF'
						if int(placar_casa) < int(placar_fora):
							time_A_fora_v += 1
							status = 'VF'
						if int(placar_casa) == int(placar_fora):
							time_A_fora_e += 1
							status = 'EF'

						time_A_fora_qtd += 1
						time_A_fora_gm += int(placar_fora)
						time_A_fora_gs += int(placar_casa)
						
					if int(placar_fora) + int(placar_casa) > 3:
						time_A_qtd_mais_35 += 1
					if int(placar_fora) + int(placar_casa) > 2:
						time_A_qtd_mais_25 += 1
					if int(placar_fora) + int(placar_casa) > 1:
						time_A_qtd_mais_15 += 1
					if int(placar_fora) + int(placar_casa) > 0:
						time_A_qtd_mais_05 += 1
					
					total_jogos_time_A += 1

					texto += f'STATUS: {status} | {data} - {campeonato} - {time_casa} ({placar_casa}) x ({placar_fora}) {time_fora}\n'

			elif (time_B == time_casa) or (time_B == time_fora):
				if total_jogos_time_B <= 15:
					if time_B == time_casa:
						if int(placar_casa) > int(placar_fora):
							time_B_casa_v += 1
							status = 'VC'
						elif int(placar_casa) < int(placar_fora):
							time_B_casa_d += 1
							status = 'DC'				
						elif int(placar_casa) == int(placar_fora):
							time_B_casa_e += 1
							status = 'EC'

						time_B_casa_qtd += 1
						time_B_casa_gm += int(placar_casa)
						time_B_casa_gs += int(placar_fora)
					
					elif time_B == time_fora:
						if int(placar_casa) > int(placar_fora):
							time_B_fora_d += 1
							status = 'DF'
						if int(placar_casa) < int(placar_fora):
							time_B_fora_v += 1
							status = 'VF'
						if int(placar_casa) == int(placar_fora):
							time_B_fora_e += 1
							status = 'EF'

						time_B_fora_qtd += 1
						time_B_fora_gm += int(placar_fora)
						time_B_fora_gs += int(placar_casa)
						
					if int(placar_fora) + int(placar_casa) > 3:
						time_B_qtd_mais_35 += 1
					if int(placar_fora) + int(placar_casa) > 2:
						time_B_qtd_mais_25 += 1
					if int(placar_fora) + int(placar_casa) > 1:
						time_B_qtd_mais_15 += 1
					if int(placar_fora) + int(placar_casa) > 0:
						time_B_qtd_mais_05 += 1
				
					total_jogos_time_B += 1

					texto += f'STATUS: {status} | {data} - {campeonato} - {time_casa} ({placar_casa}) x ({placar_fora}) {time_fora}\n'
		
		except:
			texto += f'STATUS: ERRO! - {time_A} X {time_B}\n'

	arq = f'{time_A}_{time_B}.txt'
	
	with open(arq, 'w', encoding='utf-8') as arquivo:
		arquivo.write(texto)

		# print(f'{data} - {campeonato} - {time_casa} ({placar_casa}) x ({placar_fora}) {time_fora}\nResultado para o time {time}: {resultado}')
		# input()

	dados_time_A = {'time_A_casa_qtd':[],'time_A_casa_v':[],'time_A_casa_d':[],'time_A_casa_e':[],'time_A_casa_porc_v':[],'time_A_casa_porc_d':[],'time_A_casa_porc_e':[],'time_A_casa_gm':[],'time_A_casa_gs':[], 'time_A_fora_qtd':[],'time_A_fora_v':[],'time_A_fora_d':[],'time_A_fora_e':[],'time_A_fora_porc_v':[],'time_A_fora_porc_d':[],'time_A_fora_porc_e':[],'time_A_fora_gm':[],'time_A_fora_gs':[], 'time_A_qtd_mais_05':[],'time_A_qtd_mais_15':[],'time_A_qtd_mais_25':[],'time_A_qtd_mais_35':[], 'time_A_porc_mais_05':[],'time_A_porc_mais_15':[],'time_A_porc_mais_25':[],'time_A_porc_mais_35':[]}

	dados_time_B = {'time_B_casa_qtd':[],'time_B_casa_v':[],'time_B_casa_d':[],'time_B_casa_e':[],'time_B_casa_porc_v':[],'time_B_casa_porc_d':[],'time_B_casa_porc_e':[],'time_B_casa_gm':[],'time_B_casa_gs':[], 'time_B_fora_qtd':[],'time_B_fora_v':[],'time_B_fora_d':[],'time_B_fora_e':[],'time_B_fora_porc_v':[],'time_B_fora_porc_d':[],'time_B_fora_porc_e':[],'time_B_fora_gm':[],'time_B_fora_gs':[], 'time_B_qtd_mais_05':[],'time_B_qtd_mais_15':[],'time_B_qtd_mais_25':[],'time_B_qtd_mais_35':[], 'time_B_porc_mais_05':[],'time_B_porc_mais_15':[],'time_B_porc_mais_25':[],'time_B_porc_mais_35':[]}
	
	# try:
	jogos_totais_A = time_A_casa_qtd + time_A_fora_qtd
	time_A_casa_porc_d = round((time_A_casa_d / time_A_casa_qtd) *100, 2)
	time_A_casa_porc_e = round((time_A_casa_e / time_A_casa_qtd) *100, 2)
	time_A_casa_porc_v = round((time_A_casa_v / time_A_casa_qtd) *100, 2)
	
	time_A_fora_porc_d = round((time_A_fora_d / time_A_fora_qtd) *100, 2)
	time_A_fora_porc_e = round((time_A_fora_e / time_A_fora_qtd) *100, 2)
	time_A_fora_porc_v = round((time_A_fora_v / time_A_fora_qtd) *100, 2)
	
	time_A_porc_mais_05 = round((time_A_qtd_mais_05 / jogos_totais_A) *100, 2)
	time_A_porc_mais_15 = round((time_A_qtd_mais_15 / jogos_totais_A) *100, 2)
	time_A_porc_mais_25 = round((time_A_qtd_mais_25 / jogos_totais_A) *100, 2)
	time_A_porc_mais_35 = round((time_A_qtd_mais_35 / jogos_totais_A) *100, 2)
	
	jogos_totais_B = time_B_casa_qtd + time_B_fora_qtd
	time_B_casa_porc_d = round((time_B_casa_d / time_B_casa_qtd) *100, 2)
	time_B_casa_porc_e = round((time_B_casa_e / time_B_casa_qtd) *100, 2)
	time_B_casa_porc_v = round((time_B_casa_v / time_B_casa_qtd) *100, 2)

	time_B_fora_porc_d = round((time_B_fora_d / time_B_fora_qtd) *100, 2)
	time_B_fora_porc_e = round((time_B_fora_e / time_B_fora_qtd) *100, 2)
	time_B_fora_porc_v = round((time_B_fora_v / time_B_fora_qtd) *100, 2)
	
	time_B_porc_mais_05 = round((time_B_qtd_mais_05 / jogos_totais_B) *100, 2)
	time_B_porc_mais_15 = round((time_B_qtd_mais_15 / jogos_totais_B) *100, 2)
	time_B_porc_mais_25 = round((time_B_qtd_mais_25 / jogos_totais_B) *100, 2)
	time_B_porc_mais_35 = round((time_B_qtd_mais_35 / jogos_totais_B) *100, 2)

	dados_time_A['time_A_casa_qtd'].append(time_A_casa_qtd)
	dados_time_A['time_A_casa_v'].append(time_A_casa_v)
	dados_time_A['time_A_casa_d'].append(time_A_casa_d)
	dados_time_A['time_A_casa_e'].append(time_A_casa_e)
	dados_time_A['time_A_casa_porc_v'].append(time_A_casa_porc_v)
	dados_time_A['time_A_casa_porc_d'].append(time_A_casa_porc_d)
	dados_time_A['time_A_casa_porc_e'].append(time_A_casa_porc_e)
	dados_time_A['time_A_casa_gm'].append(time_A_casa_gm)
	dados_time_A['time_A_casa_gs'].append(time_A_casa_gs)

	dados_time_A['time_A_fora_qtd'].append(time_A_fora_qtd)
	dados_time_A['time_A_fora_v'].append(time_A_fora_v)
	dados_time_A['time_A_fora_d'].append(time_A_fora_d)
	dados_time_A['time_A_fora_e'].append(time_A_fora_e)
	dados_time_A['time_A_fora_porc_v'].append(time_A_fora_porc_v)
	dados_time_A['time_A_fora_porc_d'].append(time_A_fora_porc_d)
	dados_time_A['time_A_fora_porc_e'].append(time_A_fora_porc_e)
	dados_time_A['time_A_fora_gm'].append(time_A_fora_gm)
	dados_time_A['time_A_fora_gs'].append(time_A_fora_gs)
	
	dados_time_A['time_A_qtd_mais_05'].append(time_A_qtd_mais_05)
	dados_time_A['time_A_qtd_mais_15'].append(time_A_qtd_mais_15)
	dados_time_A['time_A_qtd_mais_25'].append(time_A_qtd_mais_25)
	dados_time_A['time_A_qtd_mais_35'].append(time_A_qtd_mais_35)
	dados_time_A['time_A_porc_mais_05'].append(time_A_porc_mais_05)
	dados_time_A['time_A_porc_mais_15'].append(time_A_porc_mais_15)
	dados_time_A['time_A_porc_mais_25'].append(time_A_porc_mais_25)
	dados_time_A['time_A_porc_mais_35'].append(time_A_porc_mais_35)
	
	dados_time_B['time_B_casa_qtd'].append(time_B_casa_qtd)
	dados_time_B['time_B_casa_v'].append(time_B_casa_v)
	dados_time_B['time_B_casa_d'].append(time_B_casa_d)
	dados_time_B['time_B_casa_e'].append(time_B_casa_e)
	dados_time_B['time_B_casa_porc_v'].append(time_B_casa_porc_v)
	dados_time_B['time_B_casa_porc_d'].append(time_B_casa_porc_d)
	dados_time_B['time_B_casa_porc_e'].append(time_B_casa_porc_e)
	dados_time_B['time_B_casa_gm'].append(time_B_casa_gm)
	dados_time_B['time_B_casa_gs'].append(time_B_casa_gs)

	dados_time_B['time_B_fora_qtd'].append(time_B_fora_qtd)
	dados_time_B['time_B_fora_v'].append(time_B_fora_v)
	dados_time_B['time_B_fora_d'].append(time_B_fora_d)
	dados_time_B['time_B_fora_e'].append(time_B_fora_e)
	dados_time_B['time_B_fora_porc_v'].append(time_B_fora_porc_v)
	dados_time_B['time_B_fora_porc_d'].append(time_B_fora_porc_d)
	dados_time_B['time_B_fora_porc_e'].append(time_B_fora_porc_e)
	dados_time_B['time_B_fora_gm'].append(time_B_fora_gm)
	dados_time_B['time_B_fora_gs'].append(time_B_fora_gs)
	
	dados_time_B['time_B_qtd_mais_05'].append(time_B_qtd_mais_05)
	dados_time_B['time_B_qtd_mais_15'].append(time_B_qtd_mais_15)
	dados_time_B['time_B_qtd_mais_25'].append(time_B_qtd_mais_25)
	dados_time_B['time_B_qtd_mais_35'].append(time_B_qtd_mais_35)
	dados_time_B['time_B_porc_mais_05'].append(time_B_porc_mais_05)
	dados_time_B['time_B_porc_mais_15'].append(time_B_porc_mais_15)
	dados_time_B['time_B_porc_mais_25'].append(time_B_porc_mais_25)
	dados_time_B['time_B_porc_mais_35'].append(time_B_porc_mais_35)
	
	return dados_time_B, dados_time_A

def coletar_dados_jogos(navegador, id_jogos):
	
	for link in tqdm(id_jogos, total=len(id_jogos)):
		jogo = {'date':[],'time':[],'country':[],'league':[],'home':[],'away':[],'odd_home':[],'odd_away':[],'odd_draw':[]}
		
		data, hora, pais, liga, home, away = coletar_dados_gerais_partida(navegador, link)
		odd_away, odd_home, odd_draw = coletar_odds_pre_live_partida(navegador, link)

		print(f"{data} {hora} | {pais}: {home} x {away}") 
		
		dados_time_B, dados_time_A = coletar_dados_h2h(navegador, link, home, away)
		posicoes = coletar_classificacao_campeonato(navegador, link, home, away)
	
		jogo['date'].append(data)
		jogo['time'].append(hora)
		jogo['country'].append(pais)
		jogo['league'].append(liga)
		jogo['home'].append(home)
		jogo['away'].append(away)
		jogo['odd_away'].append(odd_away)
		jogo['odd_home'].append(odd_home)
		jogo['odd_draw'].append(odd_draw)

		jogo.update(posicoes)
		jogo.update(dados_time_A)
		jogo.update(dados_time_B)

		df_jogo = pd.DataFrame(jogo)
	
		salvar_dados_jogos_excel(df_jogo)


	# return df_jogo

def buscar_data_atual():
		
	now = datetime.now()  
	dia = repr(now.day)
	mes =  repr(now.month)    
	ano =  repr(now.year)    
	hora = repr(now.hour)
	minuto = repr(now.minute)
	
	data_hoje = f'{dia}_{mes}_{ano}-{hora}_{minuto}'
	
	return data_hoje

def salvar_dados_jogos_excel(dados_jogos):
	hoje = buscar_data_atual()
	nome_arquivo = f'flashscore_partidas-{hoje}.xlsx'
	# Verifique se o arquivo já existe
	if os.path.isfile(nome_arquivo):
			# Se o arquivo já existir, carregue-o
			wb = load_workbook(nome_arquivo)
	else:
			# Se o arquivo não existir, crie um novo
			wb = Workbook()
			# Salve o arquivo Excel
			wb.save(nome_arquivo)

	# Selecione a planilha em que deseja adicionar dados
	writer = pd.ExcelWriter(nome_arquivo, engine='openpyxl')
	writer.book = wb
	# Adicione os novos dados à planilha existente
	dados_jogos.to_excel(writer, sheet_name='Planilha1', index=False, header=False, startrow=writer.sheets['Planilha1'].max_row)
	# Salve as mudanças no arquivo Excel
	writer.save()
	



periodo = 'hoje'
# periodo = 'amanha'

navegador = abrir_pagina_flashscore()
sleep(5)
id_jogos = coletar_id_jogos(navegador, periodo)
salvar_id_jogos_txt(id_jogos)

# id_jogos = ['nP0FoSLk', 'hz3etmle', 'dtDmnWrR']
id_jogos = ['AReKzuhg']
coletar_dados_jogos(navegador, id_jogos)
# salvar_dados_jogos_excel(df_jogos)
